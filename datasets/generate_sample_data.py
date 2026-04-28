"""
Sample Data Generator for SACT Scheduler
=========================================

Generates comprehensive sample data matching all requirements from
SACT_Scheduling_System_Design.md for Velindre Cancer Centre.

Data files generated:
- patients.xlsx - Patient master data with ML features
- regimens.xlsx - Treatment regimen specifications
- sites.xlsx - Site configuration
- historical_appointments.xlsx - For ML training (2000+ records)
- staff.xlsx - Staff roster
- historical_metrics.xlsx - Performance metrics
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import random
from pathlib import Path

# Set random seed for reproducibility
np.random.seed(42)
random.seed(42)

# Output directory
OUTPUT_DIR = Path(__file__).parent / "sample_data"

# SACT v4.0 NHS code columns that must be stored as text (leading zeros, e.g. '01', '06')
# Excel/pandas coerces these to integers on save/reload without explicit text formatting.
NHS_TEXT_CODE_COLUMNS = [
    'NHS_Number_Status_Indicator_Code',  # '01'-'04'
    'Intent_Of_Treatment',               # '06', '07'
    'Treatment_Context',                 # '01'-'03'
    'Clinical_Trial',                    # '01', '02'
    'Consultant_Specialty_Code',         # '370', '800', '303'
    'End_Of_Regimen_Summary',            # '01', '06'
    'Regimen_Modification',              # 'N'=no modification (Y/N flag)
]


def save_excel_preserve_text(df: pd.DataFrame, filepath: Path, text_cols=None) -> None:
    """
    Save DataFrame to Excel using openpyxl, preserving string codes as text.
    Without this, pandas/Excel coerces '01' → 1, breaking SACT v4.0 field validation.

    Strategy: capture original string values from the DataFrame BEFORE writing
    (pandas coerces them to int during the write), then re-inject them as
    explicit string cells with '@' (text) number format after the write.
    """
    from openpyxl.utils import get_column_letter
    text_cols = set(text_cols or NHS_TEXT_CODE_COLUMNS) & set(df.columns)

    df = df.copy()

    # Capture original string values per column BEFORE pandas coerces them
    original_str_values = {}
    for col in text_cols:
        original_str_values[col] = df[col].astype(str).tolist()

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        ws = writer.sheets['Sheet1']

        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name in text_cols:
                col_letter = get_column_letter(col_idx)
                for row_offset, orig_val in enumerate(original_str_values[col_name]):
                    cell = ws[f'{col_letter}{row_offset + 2}']  # +2: row 1 = header
                    cell.number_format = '@'          # Text display format
                    cell.value = orig_val             # Original string, not Excel-coerced int
OUTPUT_DIR.mkdir(exist_ok=True)

# =============================================================================
# REFERENCE DATA - FROM DESIGN DOCUMENT
# =============================================================================

# SACT v4.0 Drug Details - Maps regimen to constituent drugs
REGIMEN_DRUGS = {
    'FOLFOX': [
        {'drug': 'Oxaliplatin', 'dose_per_m2': 85, 'unit': 'mg/m2', 'route': 'IV'},
        {'drug': 'Fluorouracil', 'dose_per_m2': 400, 'unit': 'mg/m2', 'route': 'IV'},
        {'drug': 'Folinic Acid', 'dose_per_m2': 350, 'unit': 'mg/m2', 'route': 'IV'}
    ],
    'FOLFIRI': [
        {'drug': 'Irinotecan', 'dose_per_m2': 180, 'unit': 'mg/m2', 'route': 'IV'},
        {'drug': 'Fluorouracil', 'dose_per_m2': 400, 'unit': 'mg/m2', 'route': 'IV'},
        {'drug': 'Folinic Acid', 'dose_per_m2': 350, 'unit': 'mg/m2', 'route': 'IV'}
    ],
    'RCHOP': [
        {'drug': 'Rituximab', 'dose_per_m2': 375, 'unit': 'mg/m2', 'route': 'IV'},
        {'drug': 'Cyclophosphamide', 'dose_per_m2': 750, 'unit': 'mg/m2', 'route': 'IV'},
        {'drug': 'Doxorubicin', 'dose_per_m2': 50, 'unit': 'mg/m2', 'route': 'IV'},
        {'drug': 'Vincristine', 'dose_per_m2': 1.4, 'unit': 'mg/m2', 'route': 'IV'},
        {'drug': 'Prednisolone', 'dose_per_m2': 100, 'unit': 'mg', 'route': 'PO'}
    ],
    'DOCE': [{'drug': 'Docetaxel', 'dose_per_m2': 75, 'unit': 'mg/m2', 'route': 'IV'}],
    'PACW': [{'drug': 'Paclitaxel', 'dose_per_m2': 80, 'unit': 'mg/m2', 'route': 'IV'}],
    'CARBPAC': [
        {'drug': 'Carboplatin', 'dose_per_m2': 0, 'unit': 'AUC5', 'route': 'IV'},
        {'drug': 'Paclitaxel', 'dose_per_m2': 175, 'unit': 'mg/m2', 'route': 'IV'}
    ],
    'PEMBRO': [{'drug': 'Pembrolizumab', 'dose_per_m2': 200, 'unit': 'mg', 'route': 'IV'}],
    'NIVO': [{'drug': 'Nivolumab', 'dose_per_m2': 240, 'unit': 'mg', 'route': 'IV'}],
    'TRAS': [{'drug': 'Trastuzumab', 'dose_per_m2': 6, 'unit': 'mg/kg', 'route': 'IV'}],
    'FECT': [
        {'drug': 'Fluorouracil', 'dose_per_m2': 500, 'unit': 'mg/m2', 'route': 'IV'},
        {'drug': 'Epirubicin', 'dose_per_m2': 100, 'unit': 'mg/m2', 'route': 'IV'},
        {'drug': 'Cyclophosphamide', 'dose_per_m2': 500, 'unit': 'mg/m2', 'route': 'IV'}
    ],
    'GEM': [{'drug': 'Gemcitabine', 'dose_per_m2': 1000, 'unit': 'mg/m2', 'route': 'IV'}],
    'CAPOX': [
        {'drug': 'Oxaliplatin', 'dose_per_m2': 130, 'unit': 'mg/m2', 'route': 'IV'},
        {'drug': 'Capecitabine', 'dose_per_m2': 1000, 'unit': 'mg/m2', 'route': 'PO'}
    ],
    'RITUX': [{'drug': 'Rituximab', 'dose_per_m2': 375, 'unit': 'mg/m2', 'route': 'IV'}],
    'CISE': [
        {'drug': 'Cisplatin', 'dose_per_m2': 80, 'unit': 'mg/m2', 'route': 'IV'},
        {'drug': 'Etoposide', 'dose_per_m2': 100, 'unit': 'mg/m2', 'route': 'IV'}
    ],
    'IPNIVO': [
        {'drug': 'Ipilimumab', 'dose_per_m2': 3, 'unit': 'mg/kg', 'route': 'IV'},
        {'drug': 'Nivolumab', 'dose_per_m2': 1, 'unit': 'mg/kg', 'route': 'IV'}
    ],
    'ZOLE': [{'drug': 'Zoledronic Acid', 'dose_per_m2': 4, 'unit': 'mg', 'route': 'IV'}],
    'AC': [
        {'drug': 'Doxorubicin', 'dose_per_m2': 60, 'unit': 'mg/m2', 'route': 'IV'},
        {'drug': 'Cyclophosphamide', 'dose_per_m2': 600, 'unit': 'mg/m2', 'route': 'IV'}
    ],
    'BEVA': [{'drug': 'Bevacizumab', 'dose_per_m2': 7.5, 'unit': 'mg/kg', 'route': 'IV'}],
    'PEME': [
        {'drug': 'Pemetrexed', 'dose_per_m2': 500, 'unit': 'mg/m2', 'route': 'IV'},
        {'drug': 'Carboplatin', 'dose_per_m2': 0, 'unit': 'AUC5', 'route': 'IV'}
    ],
    'VINO': [{'drug': 'Vinorelbine', 'dose_per_m2': 25, 'unit': 'mg/m2', 'route': 'IV'}],
}

# ICD-10 codes mapped to cancer types and typical regimens
ICD10_CANCER_MAP = {
    'Colorectal': {'codes': ['C18.9', 'C19', 'C20'], 'morphology': '8140/3'},
    'Lymphoma': {'codes': ['C83.3', 'C85.9'], 'morphology': '9680/3'},
    'Breast': {'codes': ['C50.9', 'C50.4'], 'morphology': '8500/3'},
    'Breast/Prostate': {'codes': ['C50.9', 'C61'], 'morphology': '8500/3'},
    'Breast/Ovarian': {'codes': ['C50.9', 'C56.9'], 'morphology': '8500/3'},
    'Ovarian/Lung': {'codes': ['C56.9', 'C34.9'], 'morphology': '8140/3'},
    'Various': {'codes': ['C34.9', 'C43.9', 'C64.9'], 'morphology': '8140/3'},
    'Pancreatic/Lung': {'codes': ['C25.9', 'C34.9'], 'morphology': '8140/3'},
    'Lung': {'codes': ['C34.9', 'C34.1'], 'morphology': '8070/3'},
    'Lung/Testicular': {'codes': ['C34.9', 'C62.9'], 'morphology': '8070/3'},
    'Melanoma': {'codes': ['C43.9', 'C43.5'], 'morphology': '8720/3'},
    'Bone Metastases': {'codes': ['C79.5'], 'morphology': '8000/3'},
    'Colorectal/Ovarian': {'codes': ['C18.9', 'C56.9'], 'morphology': '8140/3'},
    'Breast/Lung': {'codes': ['C50.9', 'C34.9'], 'morphology': '8500/3'},
}

# Treatment regimens - Using SACT standard protocol abbreviation codes
REGIMENS = {
    # Multi-day chemotherapy regimens
    'FOLFOX': {
        'name': 'FOLFOX',
        'cycle_length_days': 14,
        'treatment_days': '1,2',
        'duration_c1': 180,
        'duration_c2': 165,
        'duration_c3_plus': 150,
        'nursing_ratio': '1:3',
        'pharmacy_lead_time': 60,
        'long_infusion': True,
        'cancer_type': 'Colorectal'
    },
    'FOLFIRI': {
        'name': 'FOLFIRI',
        'cycle_length_days': 14,
        'treatment_days': '1,2',
        'duration_c1': 180,
        'duration_c2': 165,
        'duration_c3_plus': 150,
        'nursing_ratio': '1:3',
        'pharmacy_lead_time': 60,
        'long_infusion': True,
        'cancer_type': 'Colorectal'
    },
    'RCHOP': {
        'name': 'R-CHOP',
        'cycle_length_days': 21,
        'treatment_days': '1',
        'duration_c1': 360,
        'duration_c2': 300,
        'duration_c3_plus': 240,
        'nursing_ratio': '1:1',
        'pharmacy_lead_time': 90,
        'long_infusion': True,
        'cancer_type': 'Lymphoma'
    },
    'DOCE': {
        'name': 'Docetaxel',
        'cycle_length_days': 21,
        'treatment_days': '1',
        'duration_c1': 150,
        'duration_c2': 120,
        'duration_c3_plus': 90,
        'nursing_ratio': '1:4',
        'pharmacy_lead_time': 45,
        'long_infusion': False,
        'cancer_type': 'Breast/Prostate'
    },
    'PACW': {
        'name': 'Paclitaxel Weekly',
        'cycle_length_days': 7,
        'treatment_days': '1',
        'duration_c1': 180,
        'duration_c2': 150,
        'duration_c3_plus': 120,
        'nursing_ratio': '1:3',
        'pharmacy_lead_time': 45,
        'long_infusion': False,
        'cancer_type': 'Breast/Ovarian'
    },
    'CARBPAC': {
        'name': 'Carboplatin/Paclitaxel',
        'cycle_length_days': 21,
        'treatment_days': '1,8',
        'duration_c1': 240,
        'duration_c2': 210,
        'duration_c3_plus': 180,
        'nursing_ratio': '1:2',
        'pharmacy_lead_time': 60,
        'long_infusion': True,
        'cancer_type': 'Ovarian/Lung'
    },
    'PEMBRO': {
        'name': 'Pembrolizumab',
        'cycle_length_days': 21,
        'treatment_days': '1',
        'duration_c1': 60,
        'duration_c2': 45,
        'duration_c3_plus': 30,
        'nursing_ratio': '1:5',
        'pharmacy_lead_time': 30,
        'long_infusion': False,
        'cancer_type': 'Various'
    },
    'NIVO': {
        'name': 'Nivolumab',
        'cycle_length_days': 14,
        'treatment_days': '1',
        'duration_c1': 60,
        'duration_c2': 45,
        'duration_c3_plus': 30,
        'nursing_ratio': '1:5',
        'pharmacy_lead_time': 30,
        'long_infusion': False,
        'cancer_type': 'Various'
    },
    'TRAS': {
        'name': 'Trastuzumab',
        'cycle_length_days': 21,
        'treatment_days': '1',
        'duration_c1': 120,
        'duration_c2': 90,
        'duration_c3_plus': 60,
        'nursing_ratio': '1:4',
        'pharmacy_lead_time': 30,
        'long_infusion': False,
        'cancer_type': 'Breast'
    },
    'FECT': {
        'name': 'FEC-D',
        'cycle_length_days': 21,
        'treatment_days': '1',
        'duration_c1': 120,
        'duration_c2': 90,
        'duration_c3_plus': 90,
        'nursing_ratio': '1:3',
        'pharmacy_lead_time': 45,
        'long_infusion': False,
        'cancer_type': 'Breast'
    },
    'GEM': {
        'name': 'Gemcitabine',
        'cycle_length_days': 28,
        'treatment_days': '1,8,15',
        'duration_c1': 60,
        'duration_c2': 45,
        'duration_c3_plus': 30,
        'nursing_ratio': '1:5',
        'pharmacy_lead_time': 30,
        'long_infusion': False,
        'cancer_type': 'Pancreatic/Lung'
    },
    'CAPOX': {
        'name': 'CAPOX',
        'cycle_length_days': 21,
        'treatment_days': '1',
        'duration_c1': 180,
        'duration_c2': 150,
        'duration_c3_plus': 120,
        'nursing_ratio': '1:3',
        'pharmacy_lead_time': 45,
        'long_infusion': False,
        'cancer_type': 'Colorectal'
    },
    'RITUX': {
        'name': 'Rituximab Maintenance',
        'cycle_length_days': 60,
        'treatment_days': '1',
        'duration_c1': 300,
        'duration_c2': 240,
        'duration_c3_plus': 180,
        'nursing_ratio': '1:1',
        'pharmacy_lead_time': 60,
        'long_infusion': True,
        'cancer_type': 'Lymphoma'
    },
    'CISE': {
        'name': 'Cisplatin/Etoposide',
        'cycle_length_days': 21,
        'treatment_days': '1,2,3',
        'duration_c1': 300,
        'duration_c2': 270,
        'duration_c3_plus': 240,
        'nursing_ratio': '1:1',
        'pharmacy_lead_time': 90,
        'long_infusion': True,
        'cancer_type': 'Lung/Testicular'
    },
    'IPNIVO': {
        'name': 'Ipilimumab/Nivolumab',
        'cycle_length_days': 21,
        'treatment_days': '1',
        'duration_c1': 120,
        'duration_c2': 90,
        'duration_c3_plus': 60,
        'nursing_ratio': '1:2',
        'pharmacy_lead_time': 45,
        'long_infusion': False,
        'cancer_type': 'Melanoma'
    },
    'ZOLE': {
        'name': 'Zoledronic Acid',
        'cycle_length_days': 28,
        'treatment_days': '1',
        'duration_c1': 30,
        'duration_c2': 20,
        'duration_c3_plus': 15,
        'nursing_ratio': '1:6',
        'pharmacy_lead_time': 15,
        'long_infusion': False,
        'cancer_type': 'Bone Metastases'
    },
    'AC': {
        'name': 'AC (Doxorubicin/Cyclophosphamide)',
        'cycle_length_days': 21,
        'treatment_days': '1',
        'duration_c1': 90,
        'duration_c2': 75,
        'duration_c3_plus': 60,
        'nursing_ratio': '1:4',
        'pharmacy_lead_time': 45,
        'long_infusion': False,
        'cancer_type': 'Breast'
    },
    'BEVA': {
        'name': 'Bevacizumab Maintenance',
        'cycle_length_days': 14,
        'treatment_days': '1',
        'duration_c1': 90,
        'duration_c2': 60,
        'duration_c3_plus': 30,
        'nursing_ratio': '1:5',
        'pharmacy_lead_time': 30,
        'long_infusion': False,
        'cancer_type': 'Colorectal/Ovarian'
    },
    'PEME': {
        'name': 'Pemetrexed/Carboplatin',
        'cycle_length_days': 21,
        'treatment_days': '1',
        'duration_c1': 120,
        'duration_c2': 90,
        'duration_c3_plus': 75,
        'nursing_ratio': '1:3',
        'pharmacy_lead_time': 45,
        'long_infusion': False,
        'cancer_type': 'Lung'
    },
    'VINO': {
        'name': 'Vinorelbine',
        'cycle_length_days': 7,
        'treatment_days': '1',
        'duration_c1': 30,
        'duration_c2': 20,
        'duration_c3_plus': 15,
        'nursing_ratio': '1:6',
        'pharmacy_lead_time': 20,
        'long_infusion': False,
        'cancer_type': 'Breast/Lung'
    }
}

# South Wales postcodes with travel times (per Appendix A)
POSTCODES = {
    'CF10': {'area': 'Cardiff City Centre', 'distance_km': 4, 'travel_time_min': 15},
    'CF11': {'area': 'Cardiff Canton', 'distance_km': 5, 'travel_time_min': 18},
    'CF14': {'area': 'Whitchurch/Rhiwbina', 'distance_km': 0, 'travel_time_min': 5},
    'CF15': {'area': 'Radyr/Llandaff', 'distance_km': 3, 'travel_time_min': 12},
    'CF23': {'area': 'Pentwyn/Llanedeyrn', 'distance_km': 6, 'travel_time_min': 20},
    'CF24': {'area': 'Roath/Cathays', 'distance_km': 4, 'travel_time_min': 15},
    'CF3': {'area': 'Rumney/St Mellons', 'distance_km': 8, 'travel_time_min': 25},
    'CF5': {'area': 'Ely/Fairwater', 'distance_km': 6, 'travel_time_min': 22},
    'CF37': {'area': 'Pontypridd', 'distance_km': 12, 'travel_time_min': 30},
    'CF38': {'area': 'Llantrisant', 'distance_km': 14, 'travel_time_min': 35},
    'CF62': {'area': 'Barry', 'distance_km': 15, 'travel_time_min': 35},
    'CF63': {'area': 'Penarth', 'distance_km': 12, 'travel_time_min': 28},
    'CF64': {'area': 'Dinas Powys', 'distance_km': 10, 'travel_time_min': 25},
    'CF72': {'area': 'Talbot Green', 'distance_km': 10, 'travel_time_min': 25},
    'CF81': {'area': 'Bargoed', 'distance_km': 20, 'travel_time_min': 45},
    'CF82': {'area': 'Caerphilly', 'distance_km': 18, 'travel_time_min': 40},
    'CF83': {'area': 'Caerphilly North', 'distance_km': 15, 'travel_time_min': 35},
    'CF31': {'area': 'Bridgend', 'distance_km': 25, 'travel_time_min': 40},
    'CF32': {'area': 'Bridgend North', 'distance_km': 28, 'travel_time_min': 45},
    'NP10': {'area': 'Newport East', 'distance_km': 18, 'travel_time_min': 35},
    'NP19': {'area': 'Newport Central', 'distance_km': 16, 'travel_time_min': 30},
    'NP20': {'area': 'Newport West', 'distance_km': 15, 'travel_time_min': 28},
    'NP44': {'area': 'Cwmbran', 'distance_km': 20, 'travel_time_min': 40},
    'SA1': {'area': 'Swansea Central', 'distance_km': 65, 'travel_time_min': 75},
    'SA2': {'area': 'Swansea West', 'distance_km': 68, 'travel_time_min': 80},
}

# Postcode sampling weights — calibrated against pseudonymised Velindre patient
# travel data (n=5,116 patients; distances real, identifiers synthetic).
# Source file: prepare doc/Patient Data ANONYMISED.csv — contains real miles &
# minutes to each of 4 Velindre sites, with SyntheticPatientID pseudonyms.
# Real distribution: Near (<20 min) ~40%, Medium (20-45 min) ~57%, Remote (>45 min) ~3%
# Without weights, uniform sampling over 25 postcodes gives 20% Near / 72% Medium / 8% Remote
# — over-inflating remote patients vs. real Welsh geography.
#
# Weight derivation:
#   Near  (5 postcodes): 40% / 5  = 8.0 each
#   Medium (18 postcodes): 57% / 18 = 3.17 each
#   Remote (2 postcodes):  3% / 2  = 1.5 each
POSTCODE_WEIGHTS = {
    # Near (<20 min) — weight 8.0 each → collective 40%
    'CF10': 8.0, 'CF11': 8.0, 'CF14': 8.0, 'CF15': 8.0, 'CF24': 8.0,
    # Medium (20-45 min) — weight 3.17 each → collective 57%
    'CF23': 3.17, 'CF3': 3.17, 'CF5': 3.17, 'CF37': 3.17, 'CF38': 3.17,
    'CF62': 3.17, 'CF63': 3.17, 'CF64': 3.17, 'CF72': 3.17, 'CF81': 3.17,
    'CF82': 3.17, 'CF83': 3.17, 'CF31': 3.17, 'CF32': 3.17, 'NP10': 3.17,
    'NP19': 3.17, 'NP20': 3.17, 'NP44': 3.17,
    # Remote (>45 min) — weight 1.5 each → collective 3%
    'SA1': 1.5, 'SA2': 1.5,
}

# Sites configuration — Real Velindre Cancer Centre capacity
# Sources: velindre.nhs.wales, Tenovus mobile unit reports
#
# IMPORTANT: Chairs only in the day-case scheduling pool.
# The 14-bed Chemotherapy Inpatient Unit (CIU) is a separate ward
# for admitted patients and is NOT part of day-case scheduling.
#
# Long_Infusion flag replaces Requires_Bed — means patient needs
# a recliner chair or side room for comfort (4+ hour treatments),
# NOT an actual inpatient bed.
SITES = [
    {'code': 'WC', 'name': 'Velindre Whitchurch (Day Unit)', 'chairs': 19, 'hours': '08:30-18:00', 'recliners': 4, 'nurses_am': 10, 'nurses_pm': 8, 'lat': 51.5200, 'lon': -3.2100},
    {'code': 'PCH', 'name': 'Prince Charles Hospital (Macmillan Unit)', 'chairs': 11, 'hours': '09:00-17:00', 'recliners': 2, 'nurses_am': 4, 'nurses_pm': 3, 'lat': 51.7490, 'lon': -3.3780},
    {'code': 'RGH', 'name': 'Royal Glamorgan Hospital (Outreach)', 'chairs': 6, 'hours': '09:00-17:00', 'recliners': 1, 'nurses_am': 3, 'nurses_pm': 2, 'lat': 51.5728, 'lon': -3.3868},
    {'code': 'POW', 'name': 'Princess of Wales Hospital (Outreach)', 'chairs': 6, 'hours': '09:00-17:00', 'recliners': 1, 'nurses_am': 3, 'nurses_pm': 2, 'lat': 51.5040, 'lon': -3.5760},
    {'code': 'CWM', 'name': 'Cwmbran Mobile Unit (Tenovus)', 'chairs': 3, 'hours': '09:00-16:00', 'recliners': 0, 'nurses_am': 2, 'nurses_pm': 1, 'lat': 51.6530, 'lon': -3.0210},
]

# Velindre Chemotherapy Inpatient Unit (CIU) — reference only, NOT in scheduler
# 14 beds + 2 isolation cubicles, 24/7, for admitted patients
INPATIENT_UNIT = {
    'name': 'Chemotherapy Inpatient Unit (CIU)',
    'beds': 14,
    'isolation_cubicles': 2,
    'hours': '24/7',
    'ward_manager': 'Ward-level management',
    'note': 'Separate from day-case scheduling. For inpatient chemo and side-effect management (e.g., neutropenic sepsis).'
}

# Weather conditions for historical data
WEATHER_CONDITIONS = [
    ('Clear', 0.0), ('Partly Cloudy', 0.0), ('Cloudy', 0.05),
    ('Light Rain', 0.1), ('Rain', 0.2), ('Heavy Rain', 0.35),
    ('Fog', 0.15), ('Snow', 0.5), ('Ice', 0.6), ('Storm', 0.7)
]

# Welsh names
FIRST_NAMES = [
    'Rhys', 'Dylan', 'Owen', 'Gareth', 'Iwan', 'Dafydd', 'Huw', 'Geraint', 'Emyr', 'Aled',
    'Sian', 'Carys', 'Megan', 'Elin', 'Ffion', 'Catrin', 'Rhian', 'Bethan', 'Angharad', 'Gwen',
    'John', 'David', 'Michael', 'James', 'Robert', 'William', 'Richard', 'Thomas', 'Paul', 'Mark',
    'Sarah', 'Elizabeth', 'Margaret', 'Susan', 'Patricia', 'Jennifer', 'Linda', 'Barbara', 'Karen', 'Helen'
]

SURNAMES = [
    'Jones', 'Williams', 'Davies', 'Evans', 'Thomas', 'Roberts', 'Lewis', 'Morgan', 'Hughes', 'Edwards',
    'Griffiths', 'Lloyd', 'Price', 'Jenkins', 'James', 'Phillips', 'Morris', 'Owen', 'Powell', 'Rees'
]

# =============================================================================
# DATA GENERATION FUNCTIONS
# =============================================================================

_pid_rng = random.Random(999983)  # Independent RNG for patient IDs — breaks seed correlation

def generate_patient_id():
    """Generate NHS-style local patient identifier using independent RNG."""
    return f"P{_pid_rng.randint(10000, 99999)}"

def generate_nhs_number():
    """
    Generate synthetic 10-digit NHS number with valid Modulus 11 check digit.

    NHS Number format: 9 digits + 1 check digit
    Check digit algorithm: Modulus 11 (weights 10,9,8,7,6,5,4,3,2)
    Valid range: 400-799 million (England)

    These are SYNTHETIC — not real patient NHS numbers.
    """
    while True:
        # Generate 9 random digits (starting 4-7 for England range)
        first = random.randint(4, 7)
        rest = [random.randint(0, 9) for _ in range(8)]
        digits = [first] + rest

        # Calculate Modulus 11 check digit
        weights = [10, 9, 8, 7, 6, 5, 4, 3, 2]
        total = sum(d * w for d, w in zip(digits, weights))
        remainder = total % 11
        check = 11 - remainder

        if check == 11:
            check = 0
        elif check == 10:
            continue  # Invalid — regenerate

        digits.append(check)
        return ''.join(str(d) for d in digits)

def generate_regimens():
    """Generate regimens.xlsx per design doc Section 11.1"""
    records = []
    for code, reg in REGIMENS.items():
        records.append({
            'Regimen_Code': code,
            'Regimen_Name': reg['name'],
            'Cycle_Length_Days': reg['cycle_length_days'],
            'Treatment_Days': reg['treatment_days'],
            'Duration_C1': reg['duration_c1'],
            'Duration_C2': reg['duration_c2'],
            'Duration_C3_Plus': reg['duration_c3_plus'],
            'Nursing_Ratio': reg['nursing_ratio'],
            'Pharmacy_Lead_Time': reg['pharmacy_lead_time'],
            'Long_Infusion': reg['long_infusion'],
            'Cancer_Type': reg['cancer_type']
        })
    return pd.DataFrame(records)

def generate_sites():
    """Generate sites.xlsx — real Velindre Cancer Centre locations (chairs only, no beds in scheduler)"""
    records = []
    for site in SITES:
        records.append({
            'Site_Code': site['code'],
            'Site_Name': site['name'],
            'Chairs': site['chairs'],
            'Recliners': site.get('recliners', 0),  # Recliner chairs for long infusions (4+ hours)
            'Operating_Hours': site['hours'],
            'Nurses_AM': site['nurses_am'],
            'Nurses_PM': site['nurses_pm'],
            'Latitude': site['lat'],
            'Longitude': site['lon']
        })
    return pd.DataFrame(records)

def get_age_band(age):
    """Convert age to age band per PDF requirements"""
    if age < 40:
        return '<40'
    elif age < 60:
        return '40-60'
    elif age < 75:
        return '60-75'
    else:
        return '>75'


def generate_patients(n_patients=250):
    """Generate patients.xlsx per design doc Section 11.1 and PDF requirements (22 fields)"""
    # ── NHS open data calibration (CWT national 31D all-modalities) ──────
    # Replaces uniform regimen sampling with frequency-weighted sampling
    # that matches real England cancer-type distribution.
    # Importable both as a package (via flask_app.py) and as a script
    # (python datasets/generate_sample_data.py).
    _calib = None
    try:
        from datasets._nhs_calibration import load as _load_nhs_calib
        _calib = _load_nhs_calib()
    except ImportError:
        try:
            import sys as _sys
            _sys.path.insert(0, str(Path(__file__).parent))
            from _nhs_calibration import load as _load_nhs_calib
            _calib = _load_nhs_calib()
        except Exception as _exc:
            print(f"[warning] NHS calibration unavailable: {_exc}")
    except Exception as _exc:
        print(f"[warning] NHS calibration unavailable: {_exc}")

    patients = []
    regimen_codes = list(REGIMENS.keys())
    postcode_list = list(POSTCODES.keys())
    # Calibrated weights: Near 40% / Medium 57% / Remote 3%
    # (pseudonymised Velindre travel data, n=5,116; distances real, IDs synthetic).
    postcode_sample_weights = [POSTCODE_WEIGHTS[pc] for pc in postcode_list]

    # Build regimen sampling weights from the CWT-derived internal cancer_type
    # distribution.  Every regimen inherits the weight of its cancer_type; ties
    # within a cancer_type are split uniformly.  Falls back to uniform if CWT
    # data is unavailable so the generator still works offline.
    _regimen_source = 'uniform_fallback'
    if _calib and _calib.source == 'nhs_open_data' and _calib.internal_cancer_weights:
        _regimens_by_ct: dict = {}
        for code, r in REGIMENS.items():
            _regimens_by_ct.setdefault(r.get('cancer_type', 'Various'), []).append(code)
        _regimen_weights_map = {}
        for ct, weight in _calib.internal_cancer_weights.items():
            codes_in_ct = _regimens_by_ct.get(ct, [])
            if not codes_in_ct:
                continue
            per_regimen = weight / len(codes_in_ct)
            for c in codes_in_ct:
                _regimen_weights_map[c] = _regimen_weights_map.get(c, 0.0) + per_regimen
        # Any regimen with no CWT mapping gets a small residual weight so the
        # generator still produces the full regimen catalogue.
        _residual = 0.01
        for c in regimen_codes:
            _regimen_weights_map.setdefault(c, _residual)
        _regimen_sample_weights = [_regimen_weights_map[c] for c in regimen_codes]
        _regimen_source = f'cwt_{_calib.cwt_file}'
    else:
        _regimen_sample_weights = [1.0] * len(regimen_codes)
    print(f"[calibration] Regimen sampling: {_regimen_source}")

    # Contact preferences per PDF field 22
    contact_preferences = ['SMS', 'Phone', 'Email', 'Post']
    contact_weights = [0.35, 0.30, 0.25, 0.10]

    for i in range(n_patients):
        patient_id = generate_patient_id()
        regimen_code = random.choices(regimen_codes, weights=_regimen_sample_weights, k=1)[0]
        regimen = REGIMENS[regimen_code]
        postcode = random.choices(postcode_list, weights=postcode_sample_weights, k=1)[0]

        # Patient-level no-show rate.  Prior: Beta(α,β) from the calibration
        # bundle (defaults α=1.5, β=8, mean ≈ 0.158 — consistent with NHS
        # England outpatient-chemotherapy DNA rate range 7–15%).
        if _calib is not None:
            noshow_rate = np.random.beta(_calib.noshow_beta_alpha, _calib.noshow_beta_beta)
        else:
            noshow_rate = np.random.beta(1.5, 8)

        # Priority based on cancer type and randomness (PDF field 13: P1-P4)
        if regimen['cancer_type'] in ['Lymphoma', 'Testicular'] and random.random() < 0.4:
            priority = 'P1'
        elif regimen['long_infusion'] and random.random() < 0.3:
            priority = random.choice(['P1', 'P2'])
        else:
            priority = random.choices(['P1', 'P2', 'P3', 'P4'], weights=[0.1, 0.3, 0.4, 0.2])[0]

        # Site preference based on postcode
        # Site preference based on postcode — real Velindre outreach locations
        if postcode.startswith('NP'):
            site_pref = random.choice(['WC', 'RGH', 'RGH'])  # Royal Glamorgan (Llantrisant) outreach
        elif postcode in ['NP44', 'CF81', 'CF82']:
            site_pref = random.choice(['WC', 'CWM'])  # Cwmbran mobile unit
        elif postcode in ['CF31', 'CF32']:
            site_pref = random.choice(['WC', 'POW'])  # Princess of Wales, Bridgend
        elif postcode in ['CF37', 'CF38', 'CF72']:
            site_pref = random.choice(['WC', 'PCH'])  # Prince Charles Hospital, Merthyr
        else:
            site_pref = 'WC'  # Main Velindre Whitchurch

        cycle_number = random.randint(1, 12)
        total_cycles = random.randint(cycle_number, 18)

        # Age and Age Band (PDF field 14)
        age = random.randint(25, 85)
        age_band = get_age_band(age)

        # Requires 1:1 Nursing (PDF field 17) - derived from nursing ratio
        requires_1to1_nursing = regimen['nursing_ratio'] == '1:1'

        # Previous no-shows (PDF field 20) and cancellations (PDF field 21)
        prev_noshows = int(noshow_rate * random.randint(5, 20))
        prev_cancellations = random.randint(0, 5)

        # Generate SACT v4.0 compliant demographics
        first_name = random.choice(FIRST_NAMES)
        surname = random.choice(SURNAMES)
        gender_code = random.choices([1, 2, 9], weights=[0.45, 0.52, 0.03])[0]  # 1=M, 2=F, 9=Not stated
        birth_date = datetime.now() - timedelta(days=age * 365 + random.randint(0, 364))
        height_m = round(random.gauss(1.70 if gender_code == 1 else 1.62, 0.08), 2)
        weight_kg = round(random.gauss(82 if gender_code == 1 else 70, 15), 1)
        performance_status = random.choices([0, 1, 2, 3, 4], weights=[0.25, 0.35, 0.25, 0.12, 0.03])[0]

        # ICD-10 diagnosis from cancer type
        cancer_type = regimen['cancer_type']
        icd10_info = ICD10_CANCER_MAP.get(cancer_type, {'codes': ['C80.9'], 'morphology': '8000/3'})
        icd10_code = random.choice(icd10_info['codes'])

        # NHS Number (10-digit synthetic but valid format)
        nhs_number = generate_nhs_number()

        # ── Pre-compute BSA and Section 6 drug fields ─────────────────────
        _bsa = round(0.007184 * ((height_m * 100) ** 0.725) * (weight_kg ** 0.425), 3)
        _primary_drug = REGIMEN_DRUGS.get(regimen_code, [{'drug': 'Unknown', 'dose_per_m2': 0, 'unit': 'mg', 'route': 'IV'}])[0]
        _unit = _primary_drug['unit']
        _route = _primary_drug['route']
        if 'm2' in _unit:
            _daily_dose = round(_primary_drug['dose_per_m2'] * _bsa, 1)
        elif 'kg' in _unit:
            _daily_dose = round(_primary_drug['dose_per_m2'] * weight_kg, 1)
        elif 'AUC' in _unit:
            # Carboplatin AUC dosing: Calvert formula = AUC × (CrCl + 25)
            # Simplified Cockcroft-Gault: CrCl ≈ (140-age) × weight / 72 (×0.85 for female)
            _auc = float(_unit.replace('AUC', ''))
            _crcl = max(30.0, (140 - age) * weight_kg / 72 * (0.85 if gender_code == 2 else 1.0))
            _daily_dose = round(_auc * (_crcl + 25), 1)
        else:
            _daily_dose = float(_primary_drug['dose_per_m2'])

        patient = {
            # =========================================================
            # SACT v4.0 SECTION 1: LINKAGE (Mandatory)
            # =========================================================
            'NHS_Number': nhs_number,  # SACT v4.0 Mandatory
            'Local_Patient_Identifier': patient_id,  # SACT v4.0 Mandatory (internal ID)
            'NHS_Number_Status_Indicator_Code': '01',  # 01 = Number present and verified

            # =========================================================
            # SACT v4.0 SECTION 2: DEMOGRAPHICS (Required)
            # =========================================================
            'Person_Family_Name': surname,  # SACT v4.0 Required
            'Person_Given_Name': first_name,  # SACT v4.0 Required
            'Person_Birth_Date': birth_date.strftime('%Y-%m-%d'),  # SACT v4.0 Required
            'Person_Stated_Gender_Code': gender_code,  # SACT v4.0 Required (1=M, 2=F, 9=NS)
            'Patient_Postcode': f"{postcode} {random.randint(1,9)}{random.choice('ABCDEFGHJKLMNPQRSTUVWXYZ')}{random.choice('ABCDEFGHJKLMNPQRSTUVWXYZ')}",
            'Organisation_Identifier': 'RQF',  # Velindre Cancer Centre ODS code

            # =========================================================
            # SACT v4.0 SECTION 3: CLINICAL STATUS (Required)
            # =========================================================
            'Primary_Diagnosis_ICD10': icd10_code,  # SACT v4.0 Required
            'Morphology_ICD_O': icd10_info['morphology'],  # SACT v4.0 Required
            'Consultant_Specialty_Code': random.choice(['370', '800', '303']),  # Oncology specialties
            'Performance_Status': performance_status,  # SACT v4.0 Required (WHO 0-4)

            # =========================================================
            # SACT v4.0 SECTION 4: REGIMEN (Mixed mandatory/required)
            # =========================================================
            'Regimen_Code': regimen_code,  # SACT v4.0 Mandatory (protocol abbreviation)
            'Regimen_Name': regimen['name'],
            'Drug_Name': REGIMEN_DRUGS.get(regimen_code, [{'drug': 'Unknown'}])[0]['drug'],  # SACT v4.0 Mandatory §6 — primary drug
            'Intent_Of_Treatment': random.choices(['06', '07'], weights=[0.55, 0.45])[0],  # 06=Curative, 07=Non-curative
            'Treatment_Context': random.choices(['01', '02', '03'], weights=[0.15, 0.30, 0.55])[0],  # 01=Neo, 02=Adj, 03=SACT
            'Height_At_Start': height_m,  # SACT v4.0 Required (metres)
            'Weight_At_Start': weight_kg,  # SACT v4.0 Required (kg)
            'Clinical_Trial': random.choices(['01', '02'], weights=[0.08, 0.92])[0],  # 01=In trial, 02=Not
            'Chemoradiation': random.choices(['Y', 'N'], weights=[0.12, 0.88])[0],
            # Days between decision-to-treat and regimen start: calibrated
            # against CWT 62D drug-regimen compliance (Within-62-days fraction
            # ~ _calib.drug_62d_within_fraction, typically 0.64–0.66).
            # This preserves the real NHS ratio of compliant-vs-breach patients.
            'Date_Decision_To_Treat': (datetime.now() - timedelta(
                days=(random.randint(14, 60)
                      if (_calib is None or random.random() < _calib.drug_62d_within_fraction)
                      else random.randint(63, 120))
            )).strftime('%Y-%m-%d'),
            'Start_Date_Of_Regimen': (datetime.now() - timedelta(days=random.randint(0, 180))).strftime('%Y-%m-%d'),

            # =========================================================
            # SACT v4.0 SECTION 4 DERIVED: BSA (DuBois formula)
            # Stored at patient level using baseline height/weight.
            # Per-cycle BSA (with weight drift) is recalculated in historical_appointments.
            # =========================================================
            'BSA': _bsa,  # m² (pre-computed above)

            # =========================================================
            # SACT v4.0 SECTION 5: REGIMEN MODIFICATION (Cycle 1 baseline)
            # =========================================================
            'Regimen_Modification': 'N',           # N=No modification at baseline (SACT v4.0: Y/N)
            'Modification_Reason_Code': 0,        # 0=Not applicable (valid: 0-4 integer)
            'Toxicity_Grade': 0,                  # 0=No toxicity at baseline (CTCAE grade)

            # =========================================================
            # SACT v4.0 SECTION 6: DRUG DELIVERY (Cycle 1 baseline)
            # mg/m2 × BSA; mg/kg × weight; AUC: Calvert formula; flat mg: direct
            # =========================================================
            'Daily_Total_Dose': _daily_dose,
            'Unit_Of_Measurement': _unit,
            'SACT_Administration_Route': _route,
            'Cycle_Length_In_Days': REGIMENS.get(regimen_code, {}).get('cycle_length_days', 21),

            # =========================================================
            # SACT v4.0 SECTION 7: END OF REGIMEN (Cycle 1 — ongoing)
            # =========================================================
            'End_Of_Regimen_Summary': '06',       # 06=Treatment ongoing (updated at regimen end)

            # Patient_NoShow_Rate: SACT section-0 scheduling field (alias for Historical_NoShow_Rate)
            'Patient_NoShow_Rate': round(noshow_rate, 3),

            # =========================================================
            # SCHEDULING & OPERATIONAL FIELDS
            # =========================================================
            'Patient_ID': patient_id,  # Internal system ID (alias for Local_Patient_Identifier)
            'Cycle_Number': cycle_number,
            'Total_Cycles': total_cycles,
            'Priority': priority,  # P1-P4 (maps to Performance_Status for scheduling weight)
            'Postcode_District': postcode,
            'Travel_Distance_KM': POSTCODES[postcode]['distance_km'],
            'Travel_Time_Min': POSTCODES[postcode]['travel_time_min'],
            'Site_Preference': site_pref,
            'Historical_NoShow_Rate': round(noshow_rate, 3),
            'Historical_NoShow_Count': prev_noshows,
            'Previous_NoShows': prev_noshows,
            'Historical_Total_Appointments': random.randint(5, 30),
            'Days_Since_Last_Visit': random.randint(0, 60),
            'Previous_Cancellations': prev_cancellations,
            'Age': age,
            'Age_Band': age_band,
            'Has_Comorbidities': random.random() < 0.35,
            'IV_Access_Difficulty': random.random() < 0.15,
            'Requires_1to1_Nursing': requires_1to1_nursing,
            'Mobility_Issues': random.random() < 0.12,
            'Interpreter_Required': random.random() < 0.05,
            'Contact_Preference': random.choices(contact_preferences, weights=contact_weights)[0],
            'First_Name': first_name,
            'Surname': surname,
            'Phone': f"07{random.randint(100000000, 999999999)}",
            'Consultant': f"Dr. {random.choice(SURNAMES)}",
            'Notes': random.choice([
                '', '', '', '', '',
                'Allergic to latex',
                'Requires port access',
                'Previous infusion reaction - premedicate',
                'Hard IV access - experienced nurse required',
                'Anxious patient - extra support needed',
                'Hearing impairment',
                'Wheelchair user',
                'Prefers morning appointments',
                'Works - prefers afternoon'
            ]),
            'Status': random.choices(['Active', 'Completed', 'On_Hold'], weights=[0.85, 0.1, 0.05])[0],
            'Created_Date': (datetime.now() - timedelta(days=random.randint(30, 365))).strftime('%Y-%m-%d'),

            # =========================================================
            # NEW PATIENT-LEVEL FIELDS FOR ADVANCED ML MODELS
            # =========================================================

            # SURVIVAL ANALYSIS (2.2) - Baseline hazard features
            'Baseline_Risk_Score': round(noshow_rate * (1 + random.uniform(-0.2, 0.3)), 3),
            'Risk_Category': 'high' if noshow_rate > 0.2 else 'medium' if noshow_rate > 0.1 else 'low',

            # UPLIFT MODELING (2.3) - Intervention responsiveness
            'Preferred_Intervention': random.choices(
                ['sms_reminder', 'phone_call', 'email', 'none'],
                weights=[0.35, 0.25, 0.20, 0.20]
            )[0],
            'Intervention_Response_Rate': round(random.uniform(0.5, 0.95), 2),
            'Transport_Need_Score': round(POSTCODES[postcode]['distance_km'] / 70 + random.uniform(0, 0.2), 2),

            # MULTI-TASK LEARNING (3.1) - Joint prediction features
            'Expected_Duration_Variance': round(random.uniform(0.1, 0.3), 2),
            'Complexity_Score': round(random.uniform(0.3, 0.9), 2),
            'Comorbidity_Count': random.randint(0, 4) if random.random() < 0.35 else 0,

            # QUANTILE REGRESSION FOREST (3.2) - Distribution features
            'Historical_Duration_CV': round(random.uniform(0.1, 0.25), 3),  # Coefficient of variation
            'Duration_Trend': random.choices(['stable', 'increasing', 'decreasing'], weights=[0.6, 0.25, 0.15])[0],

            # =========================================================
            # NEW PATIENT-LEVEL FIELDS FOR ML MODELS (3.3, 4.1, 4.2, 4.4)
            # =========================================================

            # HIERARCHICAL BAYESIAN MODEL (3.3) - Patient-specific random effects
            'Patient_Baseline_Effect': round(np.random.normal(0, 0.15), 4),  # Baseline random effect
            'Patient_Variability': round(abs(np.random.normal(0.1, 0.03)), 3),  # Within-patient variability
            'Cluster_ID': f"C{random.randint(1, 10)}",  # Patient cluster for hierarchical grouping

            # CAUSAL INFERENCE (4.1, 4.2) - Patient-level causal features
            'Weather_Sensitivity': round(random.uniform(0.3, 1.0), 2),  # How weather affects this patient
            'Traffic_Sensitivity': round(random.uniform(0.2, 0.8), 2),  # How traffic affects this patient
            'Treatment_Eligibility': random.random() < 0.7,  # Eligible for intervention treatment

            # EVENT IMPACT MODEL (4.4) - Patient-specific event sensitivity
            'Event_Sensitivity': round(random.uniform(0.2, 1.0), 2),  # How local events affect attendance
            'Flexible_Schedule': random.random() < 0.4,  # Can reschedule easily

            # SACT v4.0 Section 4 - Line of treatment
            'Line_Of_Treatment': random.randint(1, 3),  # 1=first line, 2=second line, 3+=subsequent
        }
        patients.append(patient)

    df = pd.DataFrame(patients)
    # Shuffle to break any ordering correlation between Patient_ID and features
    df = df.sample(frac=1, random_state=42).reset_index(drop=True)
    return df

def generate_historical_appointments(patients_df, n_months=12):
    """
    Generate historical_appointments.xlsx for ML training
    Per design doc Section 11.1 and PDF requirements (22 fields)
    Minimum 2000+ records recommended
    """
    # NHS calibration bundle (same as generate_patients — affects Days_Waiting
    # distribution at appointment level).
    _calib = None
    try:
        from datasets._nhs_calibration import load as _load_nhs_calib
        _calib = _load_nhs_calib()
    except ImportError:
        try:
            import sys as _sys
            _sys.path.insert(0, str(Path(__file__).parent))
            from _nhs_calibration import load as _load_nhs_calib
            _calib = _load_nhs_calib()
        except Exception:
            pass
    except Exception:
        pass

    appointments = []
    apt_id = 100000

    start_date = datetime.now() - timedelta(days=n_months * 30)

    # Cancellation reasons per PDF field 19
    CANCELLATION_REASONS = ['Patient', 'Weather', 'Medical', 'Transport', 'Other']
    CANCELLATION_WEIGHTS = [0.35, 0.15, 0.25, 0.15, 0.10]

    # Get active patients for historical generation
    active_patients = patients_df[patients_df['Status'].isin(['Active', 'Completed'])]

    for _, patient in active_patients.iterrows():
        regimen = REGIMENS.get(patient['Regimen_Code'])
        if not regimen:
            continue

        # Generate appointments based on treatment schedule
        cycle_length = regimen['cycle_length_days']
        treatment_days = [int(d) for d in regimen['treatment_days'].split(',')]

        # Generate multiple cycles of historical appointments
        max_cycles = min(patient['Total_Cycles'], 12)
        num_cycles = random.randint(min(3, max_cycles), max(3, max_cycles))
        current_date = start_date + timedelta(days=random.randint(0, 60))

        # Get site configuration for chair assignment
        site_code = patient['Site_Preference']
        site = next((s for s in SITES if s['code'] == site_code), SITES[0])

        # CRITICAL FIX: Track CUMULATIVE no-shows and cancellations per patient
        # These represent counts BEFORE each appointment (not static values)
        cumulative_noshows = 0
        cumulative_cancellations = 0
        cumulative_appointments = 0

        for cycle in range(1, num_cycles + 1):
            for treatment_day in treatment_days:
                apt_date = current_date + timedelta(days=treatment_day - 1)

                # Skip weekends
                while apt_date.weekday() >= 5:
                    apt_date += timedelta(days=1)

                # Skip if in future
                if apt_date > datetime.now():
                    continue

                # Appointment time (8am-4pm start times)
                hour = random.choices([8, 9, 10, 11, 12, 13, 14, 15, 16],
                                      weights=[0.15, 0.2, 0.15, 0.1, 0.05, 0.1, 0.1, 0.1, 0.05])[0]
                apt_time = f"{hour:02d}:{random.choice(['00', '15', '30', '45'])}"

                # Get scheduled duration based on cycle (PDF field 7)
                if cycle == 1:
                    scheduled_duration = regimen['duration_c1']
                elif cycle == 2:
                    scheduled_duration = regimen['duration_c2']
                else:
                    scheduled_duration = regimen['duration_c3_plus']

                # Chair Number (PDF field 9)
                if regimen['long_infusion'] and site.get('recliners', 0) > 0:
                    chair_number = random.randint(1, site.get('recliners', 0))
                    chair_type = 'RECLINER'
                else:
                    chair_number = random.randint(1, site['chairs'])
                    chair_type = 'CHAIR'

                # Appointment Booked Date (PDF field 18)
                # Appointments typically booked 1-28 days in advance
                days_advance = random.randint(1, 28)
                booked_date = apt_date - timedelta(days=days_advance)

                # Weather conditions (affects no-show probability)
                weather, weather_severity = random.choices(WEATHER_CONDITIONS,
                    weights=[0.3, 0.2, 0.15, 0.1, 0.08, 0.05, 0.05, 0.03, 0.02, 0.02])[0]

                # Calculate no-show probability based on features
                base_noshow = patient['Historical_NoShow_Rate']
                day_of_week = apt_date.weekday()
                is_monday = 1 if day_of_week == 0 else 0
                is_friday = 1 if day_of_week == 4 else 0

                # Adjust no-show probability
                noshow_prob = base_noshow
                noshow_prob += weather_severity * 0.3  # Weather impact
                noshow_prob += is_monday * 0.05  # Monday effect
                noshow_prob += is_friday * 0.03  # Friday effect
                noshow_prob += (patient['Travel_Distance_KM'] / 100) * 0.1  # Distance impact
                noshow_prob += (1 if cycle == 1 else 0) * (-0.1)  # First cycle more likely to attend
                noshow_prob = max(0, min(0.6, noshow_prob))  # Cap at 60%

                # Determine attendance status (PDF field 11: Yes/No/Cancelled)
                rand_val = random.random()
                if rand_val > noshow_prob:
                    attended_status = 'Yes'
                    showed_up = True
                elif random.random() < 0.4:  # 40% of non-attendance is cancelled
                    attended_status = 'Cancelled'
                    showed_up = False
                else:
                    attended_status = 'No'
                    showed_up = False

                # Cancellation Reason (PDF field 19)
                if attended_status == 'Cancelled':
                    if weather_severity > 0.3:
                        cancellation_reason = 'Weather'
                    else:
                        cancellation_reason = random.choices(CANCELLATION_REASONS, weights=CANCELLATION_WEIGHTS)[0]
                else:
                    cancellation_reason = None

                # Calculate actual duration if showed up (PDF field 8)
                if showed_up:
                    # Add some variance to actual duration
                    variance = np.random.normal(0, scheduled_duration * 0.15)
                    actual_duration = max(scheduled_duration * 0.7,
                                         scheduled_duration + variance)
                    actual_duration = int(round(actual_duration))
                else:
                    actual_duration = None

                # Day of Week name (PDF field 12)
                day_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
                day_of_week_name = day_names[day_of_week]

                # Requires 1:1 Nursing (PDF field 17)
                requires_1to1_nursing = regimen['nursing_ratio'] == '1:1'

                # =========================================================
                # SACT v4.0: BSA (DuBois formula) and primary drug lookup
                # =========================================================
                h = float(patient.get('Height_At_Start', 1.70))
                w = float(patient.get('Weight_At_Start', 75.0))
                h_cm = h * 100  # DuBois formula requires height in cm
                bsa = round(0.007184 * (h_cm ** 0.725) * (w ** 0.425), 3)

                drugs = REGIMEN_DRUGS.get(patient['Regimen_Code'], [])
                primary_drug = drugs[0] if drugs else {'drug': 'Unknown', 'dose_per_m2': 0, 'unit': 'mg', 'route': 'IV'}
                # Fixed-dose drugs (mg or mg/kg) use flat dose; m2-based use BSA
                if primary_drug['unit'].endswith('/m2'):
                    daily_dose = round(primary_drug['dose_per_m2'] * bsa, 2)
                else:
                    daily_dose = primary_drug['dose_per_m2']

                # End_Of_Regimen_Summary: 06 = Treatment ongoing, 01 = Treatment completed
                end_of_regimen = '01' if cycle >= num_cycles else '06'

                apt_id += 1
                appointment = {
                    'Appointment_ID': f"APT{apt_id}",
                    'Patient_ID': patient['Patient_ID'],  # PDF field 1
                    'Date': apt_date.strftime('%Y-%m-%d'),  # PDF field 2
                    'Time': apt_time,
                    'Site_Code': site_code,  # PDF field 3
                    'Regimen_Code': patient['Regimen_Code'],  # PDF field 4
                    'Cycle_Number': cycle,  # PDF field 5
                    'Treatment_Day': f"Day {treatment_day}",  # PDF field 6
                    'Planned_Duration': scheduled_duration,  # PDF field 7
                    'Actual_Duration': actual_duration,  # PDF field 8
                    'Chair_Number': chair_number,  # PDF field 9
                    'Chair_Type': chair_type,
                    'Travel_Time_Min': patient['Travel_Time_Min'],  # PDF field 10
                    'Attended_Status': attended_status,  # PDF field 11 (Yes/No/Cancelled)
                    'Day_Of_Week': day_of_week_name,  # PDF field 12 (Mon-Fri)
                    'Day_Of_Week_Num': day_of_week,
                    'Priority': patient['Priority'],  # PDF field 13 (P1-P4)
                    'Age': patient['Age'],
                    'Age_Band': patient.get('Age_Band', get_age_band(patient['Age'])),  # PDF field 14
                    'Has_Comorbidities': patient['Has_Comorbidities'],  # PDF field 15
                    'IV_Access_Difficulty': patient['IV_Access_Difficulty'],  # PDF field 16
                    'Requires_1to1_Nursing': requires_1to1_nursing,  # PDF field 17
                    'Appointment_Booked_Date': booked_date.strftime('%Y-%m-%d'),  # PDF field 18
                    'Days_Booked_In_Advance': days_advance,
                    'Cancellation_Reason': cancellation_reason,  # PDF field 19
                    'Previous_NoShows': cumulative_noshows,  # PDF field 20 - CUMULATIVE before this appt
                    'Previous_Cancellations': cumulative_cancellations,  # PDF field 21 - CUMULATIVE
                    'Total_Appointments_Before': cumulative_appointments,  # For rate calculation
                    'Contact_Preference': patient.get('Contact_Preference', 'Phone'),  # PDF field 22
                    # Legacy fields for backward compatibility
                    'Showed_Up': showed_up,
                    'Scheduled_Duration': scheduled_duration,
                    'Weather_Conditions': weather,
                    'Weather_Severity': weather_severity,
                    'Patient_NoShow_Rate': patient['Historical_NoShow_Rate'],
                    'Travel_Distance_KM': patient['Travel_Distance_KM'],
                    'Is_First_Cycle': cycle == 1,
                    'Nursing_Ratio': regimen['nursing_ratio'],
                    'Long_Infusion': regimen['long_infusion'],
                    'Notes': '' if showed_up else f"{attended_status}: {cancellation_reason or 'No contact'}",

                    # =========================================================
                    # NEW FIELDS FOR ADVANCED ML MODELS (2.2, 2.3, 3.1, 3.2)
                    # =========================================================

                    # SURVIVAL ANALYSIS (2.2) - Time-to-event fields
                    'Days_To_Appointment': days_advance,  # Days until appointment (at booking)
                    'Appointment_Hour': hour,  # Hour of appointment (for hazard function)
                    'Risk_Assessment_Days': random.randint(1, 7),  # Days before when risk was assessed

                    # UPLIFT MODELING (2.3) - Intervention tracking
                    'Intervention_Type': random.choices(
                        ['none', 'sms_reminder', 'phone_call', 'transport_assistance', 'care_coordinator'],
                        weights=[0.40, 0.30, 0.15, 0.10, 0.05]
                    )[0] if random.random() < 0.6 else 'none',  # 60% received some intervention
                    'Intervention_Days_Before': random.randint(1, 7) if random.random() < 0.6 else 0,
                    'Reminder_Sent': random.random() < 0.7,  # 70% received reminder
                    'Phone_Call_Made': random.random() < 0.3,  # 30% received phone call

                    # MULTI-TASK LEARNING (3.1) - Joint prediction fields
                    'Complexity_Factor': round(random.uniform(0.3, 0.9), 2),  # Treatment complexity
                    'Comorbidity_Count': random.randint(0, 4) if patient['Has_Comorbidities'] else 0,
                    'Duration_Variance': round(abs(np.random.normal(0, 0.15)), 3),  # Historical variance

                    # QUANTILE REGRESSION FOREST (3.2) - Distribution fields
                    'Historical_Duration_Mean': scheduled_duration * (1 + np.random.normal(0, 0.1)),
                    'Historical_Duration_Std': scheduled_duration * abs(np.random.normal(0.1, 0.05)),
                    'Duration_Quantile_25': scheduled_duration * 0.85 if showed_up else None,
                    'Duration_Quantile_75': scheduled_duration * 1.15 if showed_up else None,
                    'Duration_Skewness': round(np.random.normal(0.3, 0.2), 3),  # Positive = right-skewed

                    # =========================================================
                    # NEW FIELDS FOR ML MODELS (3.3, 4.1, 4.2, 4.3, 4.4, 5.1)
                    # =========================================================

                    # HIERARCHICAL BAYESIAN MODEL (3.3) - Random effects
                    'Patient_Random_Effect': round(np.random.normal(0, 0.1), 4),  # Patient-level random effect
                    'Site_Random_Effect': round(np.random.normal(0, 0.05), 4),  # Site-level random effect
                    'Regimen_Random_Effect': round(np.random.normal(0, 0.08), 4),  # Regimen-level random effect

                    # CAUSAL INFERENCE / INSTRUMENTAL VARIABLES (4.1, 4.2)
                    # Weather -> Traffic_Delay -> No-Show causal chain
                    'Traffic_Delay_Minutes': max(0, round(
                        weather_severity * 15 +  # Weather impact on traffic
                        np.random.normal(0, 5) +  # Random variation
                        (5 if hour in [8, 9, 17, 18] else 0)  # Rush hour effect
                    , 1)),
                    'Road_Conditions': random.choices(
                        ['good', 'moderate', 'poor'],
                        weights=[0.7 - weather_severity * 0.5, 0.2 + weather_severity * 0.3, 0.1 + weather_severity * 0.2]
                    )[0],

                    # DOUBLE MACHINE LEARNING (4.3) - Treatment effect estimation
                    # Treatment indicators and confounders
                    'Treatment_Assigned': random.random() < 0.5,  # Binary treatment indicator
                    'Propensity_Score_True': round(0.5 + (patient['Historical_NoShow_Rate'] - 0.15) * 2, 3),  # True propensity

                    # EVENT IMPACT MODEL (4.4) - External events affecting appointments
                    'Local_Event_Active': random.random() < 0.15,  # 15% of days have local events
                    'Event_Type': random.choices(
                        ['none', 'sports', 'concert', 'roadworks', 'public_holiday', 'weather_warning'],
                        weights=[0.70, 0.08, 0.05, 0.10, 0.04, 0.03]
                    )[0],
                    'Event_Distance_KM': round(random.uniform(0, 20), 1) if random.random() < 0.15 else None,
                    'Event_Impact_Score': round(random.uniform(0, 1), 2) if random.random() < 0.15 else 0,
                    'Event_Start_Hour': random.randint(8, 20) if random.random() < 0.15 else None,

                    # CONFORMAL PREDICTION (5.1) - Uncertainty calibration fields
                    'Prediction_Uncertainty': round(abs(np.random.normal(0.15, 0.05)), 3),  # Base uncertainty
                    'Calibration_Set_Member': random.random() < 0.2,  # 20% for calibration
                    'Historical_Prediction_Error': round(np.random.normal(0, scheduled_duration * 0.1), 1),  # Historical errors

                    # =========================================================
                    # SACT v4.0 COMPLIANCE FIELDS (April 2026 Standard)
                    # =========================================================

                    # Demographics (SACT Section 2)
                    'Person_Stated_Gender_Code': random.choices([1, 2, 9], weights=[0.44, 0.53, 0.03])[0],  # 1=Male, 2=Female, 9=Not stated

                    # Clinical Status (SACT Section 3)
                    'Primary_Diagnosis_ICD10': random.choice([
                        'C18.9', 'C34.9', 'C50.9', 'C83.3', 'C43.9',  # Colorectal, Lung, Breast, Lymphoma, Melanoma
                        'C61', 'C56.9', 'C25.9', 'C64.9', 'C67.9'     # Prostate, Ovary, Pancreas, Kidney, Bladder
                    ]),
                    'Performance_Status': random.choices(
                        [0, 1, 2, 3, 4],
                        weights=[0.25, 0.35, 0.25, 0.12, 0.03]
                    )[0],  # WHO Performance Status 0-4

                    # Regimen (SACT Section 4) - Using SACT v4.0 codes
                    'Intent_Of_Treatment': random.choices(
                        ['06', '07'],  # 06=Curative, 07=Non-curative
                        weights=[0.55, 0.45]
                    )[0],
                    'Treatment_Context': random.choices(
                        ['01', '02', '03'],  # 01=Neoadjuvant, 02=Adjuvant, 03=SACT Only
                        weights=[0.15, 0.30, 0.55]
                    )[0],
                    'Clinical_Trial': random.choices(
                        ['01', '02'],  # 01=In trial, 02=Not in trial
                        weights=[0.08, 0.92]
                    )[0],
                    'Chemoradiation': random.choices(
                        ['Y', 'N'],
                        weights=[0.12, 0.88]
                    )[0],
                    # Historical appointment-level Date_Decision_To_Treat —
                    # calibrated to CWT 62D drug-regimen compliance (same split
                    # as the patient-level field above).
                    'Date_Decision_To_Treat': (current_date - timedelta(
                        days=(random.randint(7, 42)
                              if (_calib is None or random.random() < _calib.drug_62d_within_fraction)
                              else random.randint(63, 120))
                    )).strftime('%Y-%m-%d'),
                    'Start_Date_Of_Regimen': (current_date - timedelta(days=random.randint(0, 90))).strftime('%Y-%m-%d'),

                    # Modifications (SACT v4.0 Section 5 - Numeric codes)
                    'Regimen_Modification': random.choices(['Y', 'N'], weights=[0.15, 0.85])[0],
                    'Modification_Reason_Code': random.choices(
                        [0, 1, 2, 3, 4],
                        weights=[0.85, 0.03, 0.02, 0.05, 0.05]
                    )[0],  # SACT v4.0: 0=None, 1=Patient choice, 2=Organisational, 3=Clinical, 4=Toxicity
                    'Modification_Reason_Label': random.choices(
                        ['none', 'patient_choice', 'organisational', 'clinical_factors', 'toxicity'],
                        weights=[0.85, 0.03, 0.02, 0.05, 0.05]
                    )[0],  # Human-readable label
                    'Toxicity_Grade': random.choices(
                        [0, 1, 2, 3, 4, 5],
                        weights=[0.70, 0.10, 0.10, 0.06, 0.03, 0.01]
                    )[0],  # CTCAE v5.0 Grade 0-5 (5=death)

                    # =========================================================
                    # SACT v4.0 ALIGNMENT FIELDS (missing in previous version)
                    # Added to ensure synthetic data matches real SACT v4.0 CSV
                    # =========================================================

                    # Linkage (Section 1) — mandatory in every SACT row
                    'NHS_Number': patient.get('NHS_Number', ''),
                    'Local_Patient_Identifier': patient.get('Patient_ID', ''),
                    'NHS_Number_Status_Indicator_Code': patient.get('NHS_Number_Status_Indicator_Code', '01'),
                    'Organisation_Identifier': patient.get('Organisation_Identifier', 'RQF'),

                    # Demographics (Section 2) — mandatory in every SACT row
                    'Person_Family_Name': patient.get('Surname', ''),
                    'Person_Given_Name': patient.get('First_Name', ''),
                    'Person_Birth_Date': patient.get('Person_Birth_Date', ''),
                    'Patient_Postcode': patient.get('Patient_Postcode', ''),

                    # Clinical (Section 3)
                    'Morphology_ICD_O': patient.get('Morphology_ICD_O', '8000/3'),
                    'Consultant_Specialty_Code': patient.get('Consultant_Specialty_Code', '370'),

                    # Drug administration (Section 6)
                    'Drug_Name': primary_drug['drug'],
                    'Daily_Total_Dose': daily_dose,
                    'Unit_Of_Measurement': primary_drug['unit'],
                    'SACT_Administration_Route': primary_drug['route'],

                    # Regimen details (Section 4)
                    'Cycle_Length_In_Days': regimen['cycle_length_days'],
                    'BSA': bsa,
                    'Height_At_Start': h,
                    'Weight_At_Start': w,
                    'Line_Of_Treatment': patient.get('Line_Of_Treatment', random.randint(1, 3)),

                    # Outcome (Section 7)
                    'End_Of_Regimen_Summary': end_of_regimen,
                }
                appointments.append(appointment)

                # UPDATE CUMULATIVE COUNTS after this appointment
                cumulative_appointments += 1
                if attended_status == 'No':
                    cumulative_noshows += 1
                elif attended_status == 'Cancelled':
                    cumulative_cancellations += 1

            # Move to next cycle
            current_date += timedelta(days=cycle_length)

    df = pd.DataFrame(appointments)
    df = df.sort_values('Date')
    return df

def generate_future_appointments(patients_df, n_days=30):
    """Generate scheduled appointments for the next n_days"""
    appointments = []
    apt_id = 500000

    start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    active_patients = patients_df[patients_df['Status'] == 'Active']

    chair_end_times = {}  # Track chair availability: (site_code, chair) -> next_available_minute

    for day_offset in range(n_days):
        current_date = start_date + timedelta(days=day_offset)

        # Skip weekends
        if current_date.weekday() >= 5:
            continue

        # Select patients for this day (30-50 per day)
        n_appointments = random.randint(30, 50)
        day_patients = active_patients.sample(n=min(n_appointments, len(active_patients)))

        # Reset all chairs for this new day
        for s in SITES:
            open_min = int(s['hours'].split('-')[0].split(':')[0]) * 60
            for ch in range(1, s['chairs'] + 1):
                chair_end_times[(s['code'], ch)] = open_min
            for b in range(1, s.get('recliners', 0) + 1):
                chair_end_times[(s['code'], f'REC{b}')] = open_min

        # Assign each patient a preferred arrival time spread across the day
        # This creates a realistic schedule with appointments throughout the day
        day_patients_list = list(day_patients.iterrows())
        random.shuffle(day_patients_list)

        for idx, (_, patient) in enumerate(day_patients_list):
            regimen = REGIMENS.get(patient['Regimen_Code'])
            if not regimen:
                continue

            cycle = patient['Cycle_Number']
            if cycle == 1:
                duration = regimen['duration_c1']
            elif cycle == 2:
                duration = regimen['duration_c2']
            else:
                duration = regimen['duration_c3_plus']

            site_code = patient['Site_Preference']
            site = next((s for s in SITES if s['code'] == site_code), SITES[0])
            open_min = int(site['hours'].split('-')[0].split(':')[0]) * 60
            close_min = int(site['hours'].split('-')[1].split(':')[0]) * 60

            requires_1to1 = regimen['nursing_ratio'] == '1:1'

            # Patient's preferred start time — spread across the day
            # Weight towards morning but with realistic afternoon slots
            preferred_slots = []
            for h in range(open_min // 60, close_min // 60):
                for m in [0, 15, 30, 45]:
                    slot = h * 60 + m
                    if slot + duration <= close_min:
                        # Weight: morning preferred, lunch dip, afternoon recovery
                        if h < 10:
                            weight = 3.0
                        elif h < 12:
                            weight = 2.0
                        elif h < 13:
                            weight = 0.5  # Lunch
                        elif h < 15:
                            weight = 2.0
                        else:
                            weight = 1.0
                        preferred_slots.append((slot, weight))

            if not preferred_slots:
                continue

            # Pick a preferred time
            slots, weights = zip(*preferred_slots)
            total_w = sum(weights)
            weights = [w / total_w for w in weights]
            preferred_start = random.choices(slots, weights=weights)[0]

            # Find the chair with availability closest to preferred time
            if regimen['long_infusion'] and site.get('recliners', 0) > 0:
                chair_keys = [(site_code, f'REC{b}') for b in range(1, site.get('recliners', 0) + 1)]
            else:
                chair_keys = [(site_code, ch) for ch in range(1, site['chairs'] + 1)]

            best_key = None
            best_start = None
            best_distance = 99999  # Distance from preferred time

            for key in chair_keys:
                avail_at = chair_end_times.get(key, open_min)
                # Actual start is max of chair availability and preferred time
                actual_start = max(avail_at, preferred_start)
                if actual_start + duration <= close_min:
                    distance = abs(actual_start - preferred_start)
                    if distance < best_distance:
                        best_distance = distance
                        best_start = actual_start
                        best_key = key

            if best_key is None:
                continue

            start_min_val = best_start
            _, best_chair = best_key

            # Update chair availability (add 15 min buffer)
            chair_end_times[best_key] = start_min_val + duration + 15

            # Format times
            apt_time = f"{start_min_val // 60:02d}:{start_min_val % 60:02d}"
            end_total_min = start_min_val + duration
            end_time = f"{end_total_min // 60:02d}:{end_total_min % 60:02d}"

            # Build chair_id and chair_number
            # Beds get chair numbers AFTER regular chairs (e.g., chairs 1-12, beds 13-16)
            if isinstance(best_chair, str) and best_chair.startswith('REC'):
                chair_id = f"{site_code}-{best_chair}"
                bed_num = int(best_chair.replace('REC', ''))
                chair_number = site['chairs'] + bed_num  # e.g., 12 chairs + BED1 = chair 13
            else:
                chair_id = f"{site_code}-C{best_chair}"
                chair_number = int(best_chair)

            # Assign nurse
            start_hour = start_min_val // 60
            nurse_id = f"STF{1001 + hash(f'{site_code}{current_date}{start_hour}') % 20}"
            nurse_name = f"{random.choice(FIRST_NAMES)} {random.choice(SURNAMES)}"

            # Day in cycle
            treatment_days_list = [int(d) for d in regimen['treatment_days'].split(',')]
            day_in_cycle = random.choice(treatment_days_list)

            apt_id += 1
            appointment = {
                'Appointment_ID': f"APT{apt_id}",
                # ── SACT v4.0 Section 1: Patient Identifier ─────────────────
                'Local_Patient_Identifier': patient.get('Local_Patient_Identifier', patient['Patient_ID']),
                'NHS_Number': patient.get('NHS_Number', ''),
                'NHS_Number_Status_Indicator_Code': patient.get('NHS_Number_Status_Indicator_Code', '01'),
                # ── SACT v4.0 Section 2: Demographics ────────────────────────
                'Person_Family_Name': patient.get('Person_Family_Name', ''),
                'Person_Given_Name': patient.get('Person_Given_Name', ''),
                'Person_Birth_Date': patient.get('Person_Birth_Date', ''),
                'Person_Stated_Gender_Code': patient.get('Person_Stated_Gender_Code', 1),
                'Patient_Postcode': patient.get('Patient_Postcode', ''),
                'Organisation_Identifier': patient.get('Organisation_Identifier', ''),
                # ── SACT v4.0 Section 3: Diagnosis ───────────────────────────
                'Primary_Diagnosis_ICD10': patient.get('Primary_Diagnosis_ICD10', ''),
                'Morphology_ICD_O': patient.get('Morphology_ICD_O', ''),
                'Performance_Status': patient.get('Performance_Status', 1),
                'Consultant_Specialty_Code': patient.get('Consultant_Specialty_Code', '800'),
                # ── SACT v4.0 Section 4: Regimen ─────────────────────────────
                'Regimen_Code': patient['Regimen_Code'],
                'Intent_Of_Treatment': patient.get('Intent_Of_Treatment', '06'),
                'Treatment_Context': patient.get('Treatment_Context', '01'),
                'Start_Date_Of_Regimen': patient.get('Start_Date_Of_Regimen', ''),
                'Date_Decision_To_Treat': patient.get('Date_Decision_To_Treat', ''),
                'Height_At_Start': patient.get('Height_At_Start', ''),
                'Weight_At_Start': patient.get('Weight_At_Start', ''),
                'BSA': patient.get('BSA', ''),
                'Line_Of_Treatment': patient.get('Line_Of_Treatment', 1),
                'Clinical_Trial': patient.get('Clinical_Trial', '02'),
                'Chemoradiation': patient.get('Chemoradiation', 'N'),
                # ── SACT v4.0 Section 5: Modification ────────────────────────
                'Regimen_Modification': patient.get('Regimen_Modification', 'N'),
                'Modification_Reason_Code': patient.get('Modification_Reason_Code', 0),
                'Toxicity_Grade': patient.get('Toxicity_Grade', 0),
                # ── SACT v4.0 Section 6: Drug Delivery ───────────────────────
                'Drug_Name': patient.get('Drug_Name', ''),
                'Daily_Total_Dose': patient.get('Daily_Total_Dose', ''),
                'Unit_Of_Measurement': patient.get('Unit_Of_Measurement', ''),
                'SACT_Administration_Route': patient.get('SACT_Administration_Route', 'IV'),
                'Cycle_Length_In_Days': patient.get('Cycle_Length_In_Days', 21),
                # ── SACT v4.0 Section 7: End of Regimen ──────────────────────
                'End_Of_Regimen_Summary': patient.get('End_Of_Regimen_Summary', '06'),
                # ── Scheduling / Operational ──────────────────────────────────
                'Patient_ID': patient['Patient_ID'],
                'Patient_Name': f"{patient['Person_Given_Name']} {patient['Person_Family_Name']}",
                'Date': current_date.strftime('%Y-%m-%d'),
                'Start_Time': apt_time,
                'End_Time': end_time,
                'Duration_Minutes': duration,
                'Site_Code': site_code,
                'Site_Name': site['name'],
                'Chair_ID': chair_id,
                'Chair_Number': chair_number,
                'Regimen_Name': regimen['name'],
                'Cancer_Type': regimen['cancer_type'],
                'Cycle_Number': patient['Cycle_Number'],
                'Total_Cycles': patient['Total_Cycles'],
                'Day_In_Cycle': day_in_cycle,
                'Priority': patient['Priority'],
                'Long_Infusion': regimen['long_infusion'],
                'Requires_1to1_Nursing': requires_1to1,
                'Nursing_Ratio': regimen['nursing_ratio'],
                'Nurse_ID': nurse_id,
                'Nurse_Name': nurse_name,
                'Status': 'Scheduled',
                'Consultant': patient['Consultant'],
                'Notes': patient.get('Notes', ''),
                'Pre_Meds_Required': random.random() < 0.25,
                'Bloods_Required': random.random() < 0.35,
                'ML_NoShow_Probability': round(patient['Historical_NoShow_Rate'] * random.uniform(0.8, 1.3), 3),
                # ── SACT section-0 scheduling fields ──────────────────────────
                'Planned_Duration': duration,               # = Duration_Minutes (SACT name)
                'Actual_Duration': None,                    # NULL until appointment completes
                'Attended_Status': None,                     # NULL — not yet attended (future appointment)
                'Travel_Distance_KM': patient.get('Travel_Distance_KM', 0),
                'Travel_Time_Min': patient.get('Travel_Time_Min', 0),
                'Weather_Severity': round(random.choices([0.0, 0.25, 0.5, 0.75, 1.0], weights=[0.60, 0.15, 0.12, 0.08, 0.05])[0], 2),
                'Patient_NoShow_Rate': round(patient['Historical_NoShow_Rate'], 3),
                'Created_Date': datetime.now().strftime('%Y-%m-%d'),
                'Created_By': 'System'
            }
            appointments.append(appointment)

    return pd.DataFrame(appointments)

def generate_staff():
    """
    Generate staff roster with NHS-standard banding.

    NHS Agenda for Change pay bands:
        Band 5: Newly qualified nurse
        Band 6: Specialist/experienced nurse (SACT trained)
        Band 7: Advanced practitioner / team lead
        Band 8a: Nurse manager / consultant pharmacist
    """
    staff = []
    staff_id = 1000

    # (Role, Count, NHS Band, SACT-trained, IV-certified)
    roles = [
        ('SACT Nurse', 20, 'Band 6', True, True),
        ('Senior SACT Nurse', 6, 'Band 7', True, True),
        ('Nurse Manager', 2, 'Band 8a', True, True),
        ('Pharmacist', 4, 'Band 7', False, False),
        ('Pharmacy Technician', 6, 'Band 5', False, False),
        ('Consultant Oncologist', 8, 'Consultant', False, False),
        ('Registrar', 4, 'ST3-ST8', False, False),
        ('Healthcare Assistant', 10, 'Band 3', False, False),
        ('Admin Coordinator', 5, 'Band 4', False, False),
        ('Receptionist', 3, 'Band 3', False, False),
    ]

    for role, count, band, chemo_trained, iv_cert in roles:
        for i in range(count):
            staff_id += 1
            site = random.choice([s['code'] for s in SITES])
            first_name = random.choice(FIRST_NAMES)
            surname = random.choice(SURNAMES)

            staff.append({
                'Staff_ID': f"STF{staff_id}",
                'First_Name': first_name,
                'Surname': surname,
                'Name': f"{first_name} {surname}",
                'Role': role,
                'Band': band,  # NHS Agenda for Change band
                'Site_Code': site,
                'Email': f"staff{staff_id}@velindre.nhs.wales",
                'Phone_Ext': f"{random.randint(1000, 9999)}",
                'SACT_Trained': chemo_trained,
                'IV_Certified': iv_cert,
                'Max_Patients_Per_Shift': {
                    'SACT Nurse': 4, 'Senior SACT Nurse': 3,
                    'Nurse Manager': 2, 'Healthcare Assistant': 6
                }.get(role, 0),
                'Can_Administer_1to1': band in ['Band 6', 'Band 7', 'Band 8a'] and chemo_trained,
                'Start_Date': (datetime.now() - timedelta(days=random.randint(30, 3650))).strftime('%Y-%m-%d'),
                'Active': True
            })

    return pd.DataFrame(staff)

def generate_historical_metrics(n_days=180):
    """Generate historical performance metrics"""
    metrics = []
    start_date = datetime.now() - timedelta(days=n_days)

    for day_offset in range(n_days):
        current_date = start_date + timedelta(days=day_offset)

        if current_date.weekday() >= 5:
            continue

        for site in SITES:
            total_appointments = random.randint(15, 35) if site['code'] == 'WC' else random.randint(5, 15)
            noshow_rate = random.uniform(0.08, 0.18)
            noshows = int(total_appointments * noshow_rate)
            completed = total_appointments - noshows - random.randint(0, 2)

            metrics.append({
                'Date': current_date.strftime('%Y-%m-%d'),
                'Site_Code': site['code'],
                'Total_Appointments': total_appointments,
                'Completed': completed,
                'No_Shows': noshows,
                'Cancellations': random.randint(0, 3),
                'Urgent_Additions': random.randint(0, 2),
                'Average_Wait_Time': random.randint(5, 25),
                'Chair_Utilization': round(random.uniform(0.65, 0.92), 2),
                'Recliner_Utilization': round(random.uniform(0.55, 0.88), 2) if site.get('recliners', 0) > 0 else 0,
                'Average_Duration_Variance': random.randint(-15, 25),
                'Staff_Utilization': round(random.uniform(0.70, 0.95), 2),
                'Patient_Satisfaction': round(random.uniform(4.0, 5.0), 1),
                'Incidents': random.randint(0, 1),
                'Weather_Impact': random.choice(['None', 'None', 'None', 'Low', 'Medium', 'High'])
            })

    return pd.DataFrame(metrics)

# =============================================================================
# MAIN EXECUTION
# =============================================================================

if __name__ == "__main__":
    print("=" * 70)
    print("SACT Scheduler - Comprehensive Sample Data Generator")
    print("Based on SACT_Scheduling_System_Design.md specifications")
    print("=" * 70)

    # Generate regimens first
    print("\n1. Generating regimens.xlsx...")
    regimens_df = generate_regimens()
    regimens_df.to_excel(OUTPUT_DIR / "regimens.xlsx", index=False)
    print(f"   Created {len(regimens_df)} treatment regimens")

    # Generate sites
    print("\n2. Generating sites.xlsx...")
    sites_df = generate_sites()
    sites_df.to_excel(OUTPUT_DIR / "sites.xlsx", index=False)
    print(f"   Created {len(sites_df)} treatment sites")

    # Generate patients
    print("\n3. Generating patients.xlsx...")
    patients_df = generate_patients(250)
    save_excel_preserve_text(patients_df, OUTPUT_DIR / "patients.xlsx")
    print(f"   Created {len(patients_df)} patients with ML features")

    # Generate historical appointments (for ML training)
    print("\n4. Generating historical_appointments.xlsx (for ML training)...")
    print("   This may take a moment...")
    historical_df = generate_historical_appointments(patients_df, n_months=12)
    save_excel_preserve_text(historical_df, OUTPUT_DIR / "historical_appointments.xlsx")
    noshow_count = (~historical_df['Showed_Up']).sum()
    print(f"   Created {len(historical_df)} historical appointments")
    print(f"   No-shows: {noshow_count} ({noshow_count/len(historical_df)*100:.1f}%)")

    # Generate future appointments
    print("\n5. Generating appointments.xlsx (scheduled)...")
    appointments_df = generate_future_appointments(patients_df, n_days=30)
    save_excel_preserve_text(appointments_df, OUTPUT_DIR / "appointments.xlsx")
    print(f"   Created {len(appointments_df)} scheduled appointments")

    # Generate staff
    print("\n6. Generating staff.xlsx...")
    staff_df = generate_staff()
    staff_df.to_excel(OUTPUT_DIR / "staff.xlsx", index=False)
    print(f"   Created {len(staff_df)} staff members")

    # Generate historical metrics
    print("\n7. Generating historical_metrics.xlsx...")
    metrics_df = generate_historical_metrics(180)
    metrics_df.to_excel(OUTPUT_DIR / "historical_metrics.xlsx", index=False)
    print(f"   Created {len(metrics_df)} daily metrics records")

    # Summary
    print("\n" + "=" * 70)
    print("DATA GENERATION COMPLETE")
    print("=" * 70)
    print(f"\nOutput directory: {OUTPUT_DIR}")
    print("\nFiles created (per design document specifications):")
    print(f"  - regimens.xlsx              ({len(regimens_df)} regimens)")
    print(f"  - sites.xlsx                 ({len(sites_df)} sites)")
    print(f"  - patients.xlsx              ({len(patients_df)} patients)")
    print(f"  - historical_appointments.xlsx ({len(historical_df)} records for ML training)")
    print(f"  - appointments.xlsx          ({len(appointments_df)} scheduled)")
    print(f"  - staff.xlsx                 ({len(staff_df)} staff)")
    print(f"  - historical_metrics.xlsx    ({len(metrics_df)} daily records)")

    print("\nKey ML Training Features in historical_appointments.xlsx:")
    print("  - Showed_Up (boolean for no-show prediction)")
    print("  - Actual_Duration (for duration prediction)")
    print("  - Weather_Conditions & Weather_Severity")
    print("  - Patient_NoShow_Rate, Travel_Distance_KM")
    print("  - Day_Of_Week, Is_First_Cycle, Has_Comorbidities")
    print("  - Nursing_Ratio, Long_Infusion, Priority")

    print("\nAdvanced ML Model Features:")
    print("  2.2 Survival Analysis: Days_To_Appointment, Appointment_Hour, Risk_Assessment_Days")
    print("  2.3 Uplift Modeling: Intervention_Type, Reminder_Sent, Phone_Call_Made")
    print("  3.1 Multi-Task: Complexity_Factor, Comorbidity_Count, Duration_Variance")
    print("  3.2 Quantile Forest: Historical_Duration_Mean/Std, Duration_Quantile_25/75")
    print("  3.3 Hierarchical: Patient_Random_Effect, Site_Random_Effect, Regimen_Random_Effect")
    print("  4.1-4.2 Causal/IV: Traffic_Delay_Minutes, Road_Conditions, Weather_Severity")
    print("  4.3 DML: Treatment_Assigned, Propensity_Score_True")
    print("  4.4 Event Impact: Local_Event_Active, Event_Type, Event_Distance_KM, Event_Impact_Score")
    print("  5.1 Conformal: Prediction_Uncertainty, Calibration_Set_Member, Historical_Prediction_Error")
    print("  SACT v4.0: Gender_Code, ICD10_Diagnosis, Performance_Status, Intent_Of_Treatment")
    print("             Treatment_Context, Clinical_Trial, Date_Decision_To_Treat, Toxicity_Grade")
